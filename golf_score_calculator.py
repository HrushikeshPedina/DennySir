#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import openpyxl
from openpyxl import Workbook
import pandas as pd
import tempfile
import os

# ----------- Golf Logic (System 36) ------------
def calculate_system_36_9hole(pars, scores):
    points = []
    total_points = 0
    gross_score = sum(scores)

    for score, par in zip(scores, pars):
        if score <= par:
            point = 2
        elif score == par + 1:
            point = 1
        else:
            point = 0
        points.append(point)
        total_points += point

    handicap = 18 - total_points
    net_score = gross_score - handicap

    return {
        "gross_score": gross_score,
        "total_points": total_points,
        "handicap": handicap,
        "net_score": net_score,
        "points": points
    }

def process_scorecard(input_file):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active

    pars = [ws_in.cell(row=i, column=2).value for i in range(2, 11)]
    player_names = [ws_in.cell(row=1, column=col).value for col in range(3, ws_in.max_column + 1)]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Player Scores"

    current_row = 1
    comparison_data = []

    for idx, col in enumerate(range(3, ws_in.max_column + 1)):
        name = player_names[idx]
        scores = [ws_in.cell(row=i, column=col).value for i in range(2, 11)]
        result = calculate_system_36_9hole(pars, scores)

        ws_out.cell(row=current_row, column=1, value=name)
        current_row += 1

        headers = ["Hole", "Par", "Score", "System 36 Points"]
        for j, header in enumerate(headers, start=1):
            ws_out.cell(row=current_row, column=j, value=header)
        current_row += 1

        for i in range(9):
            ws_out.cell(row=current_row, column=1, value=i + 1)
            ws_out.cell(row=current_row, column=2, value=pars[i])
            ws_out.cell(row=current_row, column=3, value=scores[i])
            ws_out.cell(row=current_row, column=4, value=result["points"][i])
            current_row += 1

        ws_out.cell(row=current_row, column=1, value="Gross Score")
        ws_out.cell(row=current_row, column=2, value=result["gross_score"])
        current_row += 1

        ws_out.cell(row=current_row, column=1, value="System 36 Points")
        ws_out.cell(row=current_row, column=2, value=result["total_points"])
        current_row += 1

        ws_out.cell(row=current_row, column=1, value="Handicap (System 36)")
        ws_out.cell(row=current_row, column=2, value=result["handicap"])
        current_row += 1

        ws_out.cell(row=current_row, column=1, value="Net Score")
        ws_out.cell(row=current_row, column=2, value=result["net_score"])
        current_row += 2

        comparison_data.append({
            "name": name,
            "gross": result["gross_score"],
            "points": result["total_points"],
            "handicap": result["handicap"],
            "net": result["net_score"]
        })

    comparison_data.sort(key=lambda x: x["net"])

    ws_out.cell(row=current_row, column=1, value="Final Comparison")
    current_row += 1

    ws_out.append(["Player", "Gross Score", "System 36 Points", "Handicap", "Net Score"])
    for player in comparison_data:
        ws_out.append([
            player["name"],
            player["gross"],
            player["points"],
            player["handicap"],
            player["net"]
        ])

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb_out.save(tmp_file.name)
    return tmp_file.name, pd.DataFrame(comparison_data)

# ----------- Streamlit App UI ------------
st.set_page_config(page_title="Golf Score Calculator", page_icon="⛳", layout="centered")

st.markdown("""
    <style>
    .main {
        background-color: #f0f9f5;
    }
    .block-container {
        padding-top: 2rem;
    }
    </style>
    """, unsafe_allow_html=True)

st.title("⛳ Golf Score Calculator - System 36 (9 Holes)")
st.markdown("Upload your golf scorecard in Excel format to calculate scores using the **System 36** method for 9 holes.")

uploaded_file = st.file_uploader("📁 Upload Excel File (.xlsx)", type=["xlsx"])

if uploaded_file:
    with st.spinner("Processing your scorecard..."):
        output_file, comparison_df = process_scorecard(uploaded_file)
    st.success("Scorecard processed successfully!")

    st.markdown("### 🏆 Final Standings")
    st.dataframe(comparison_df)

    with open(output_file, "rb") as f:
        st.download_button("📥 Download Processed Excel", f, file_name="Processed_Golf_Scores.xlsx")

else:
    st.info("Please upload a valid scorecard Excel file to begin.")

